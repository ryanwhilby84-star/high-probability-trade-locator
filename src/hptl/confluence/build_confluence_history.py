from __future__ import annotations

from datetime import datetime
from dataclasses import replace
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from hptl.cot.exporter import _calculate_cot_scores
from hptl.cot.downloader import download_latest_cot
from hptl.cot.parser import parse_cot_file
from hptl.config import get_settings

EXPORT_DIR = Path("data/exports")
PROCESSED_DIR = Path("data/processed")
TARGET_ALIASES = {
    "NASDAQ / NQ": ["NASDAQ MINI", "E-MINI NASDAQ", "NASDAQ-100", "NASDAQ 100", "NASDAQ 100 STOCK INDEX"],
    "S&P 500 / ES": ["E-MINI S&P 500", "S&P 500 STOCK INDEX", "S&P 500 CONSOLIDATED", "SP 500"],
    "Dow / YM / DJIA / S30": ["DOW JONES", "DJIA", "E-MINI DOW", "MINI DOW", "DOW JONES U.S. INDEX"],
    "Gold / GC": ["GOLD -", "GOLD"],
    "Silver / SI": ["SILVER -", "SILVER"],
    "Copper / HG": ["COPPER-GRADE #1", "COPPER"],
    "Crude Oil / CL": ["CRUDE OIL, LIGHT SWEET", "CRUDE OIL"],
    "Natural Gas / NG": ["NATURAL GAS"],
    "Corn / ZC": ["CORN -"],
    "Soybeans / ZS": ["SOYBEANS -"],
    "Wheat / ZW": ["WHEAT -"],
    "Coffee / KC": ["COFFEE C -", "COFFEE"],
    "Cocoa / CC": ["COCOA -", "COCOA"],
}
TARGET_MARKETS = list(TARGET_ALIASES.keys())
HISTORY_START_DATE = pd.Timestamp("2024-05-06")
HISTORY_END_DATE = pd.Timestamp("2026-05-06")


def _normalize_column_name(col: str) -> str:
    return " ".join(str(col).strip().lower().replace("_", " ").split())


def _find_column(df: pd.DataFrame, *aliases: str) -> str | None:
    normalized = {_normalize_column_name(col): col for col in df.columns}
    for alias in aliases:
        col = normalized.get(_normalize_column_name(alias))
        if col is not None:
            return col
    return None


def _clean_bias(value: Any) -> str:
    text = str(value).strip().lower()
    if "bull" in text or text in {"long", "buy"}:
        return "Bullish"
    if "bear" in text or text in {"short", "sell"}:
        return "Bearish"
    return "Neutral"


def _clean_strength(value: Any) -> str:
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return "Unknown"
    return text


def _macro_alignment_adjustment(score: float) -> int:
    if score <= 2:
        return 1
    if score <= 5:
        return 2
    if score <= 7:
        return 3
    return 4


def _macro_signal_label(signal: str) -> str:
    if signal == "risk_on":
        return "risk-on"
    if signal == "risk_off":
        return "risk-off"
    return "neutral"


def _strength_from_score(score: float) -> str:
    if score >= 8:
        return "Very Strong"
    if score >= 6:
        return "Strong"
    if score >= 3:
        return "Moderate"
    return "Weak"


def _discover_cot_files() -> list[Path]:
    files = sorted(PROCESSED_DIR.glob("cot_cleaned_*.csv"), key=lambda p: p.stat().st_mtime)
    if not files:
        raise FileNotFoundError("No COT history inputs found. Expected data/processed/cot_cleaned_*.csv")
    return files


def _ensure_cot_backfill(start_year: int = 2023) -> None:
    """Safely ensure annual COT cleaned files exist from start_year to current year."""
    current_year = datetime.utcnow().year
    settings = get_settings()
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

    for year in range(start_year, current_year + 1):
        existing = list(PROCESSED_DIR.glob(f"cot_cleaned_{year}*.csv"))
        if existing:
            continue
        year_settings = replace(settings, cot_year=year)
        print(f"Backfill: downloading COT annual file for {year}")
        try:
            download = download_latest_cot(year_settings)
            parsed = parse_cot_file(download.raw_file_path)
            cleaned_path = PROCESSED_DIR / f"cot_cleaned_{year}_backfill.csv"
            parsed.to_csv(cleaned_path, index=False)
            print(f"Backfill: wrote {cleaned_path}")
        except Exception as exc:
            print(f"WARNING: Backfill failed for {year}: {exc}")
            print("Continuing with already available local COT files.")


def _discover_macro_files() -> tuple[list[Path], str, str]:
    history_files = sorted(EXPORT_DIR.glob("macro_history_*.xlsx"), key=lambda p: p.stat().st_mtime)
    if history_files:
        return history_files, "Macro_History", "macro_history_*.xlsx"

    output_files = sorted(EXPORT_DIR.glob("macro_output_*.xlsx"), key=lambda p: p.stat().st_mtime)
    if output_files:
        return output_files, "Macro_Dashboard", "macro_output_*.xlsx"

    raise FileNotFoundError(
        "No macro inputs found. Expected data/exports/macro_history_*.xlsx (preferred) or macro_output_*.xlsx"
    )


def _normalize_market_text(v: str) -> str:
    return " ".join(v.upper().replace("_", " ").replace("/", " ").replace("-", " ").split())


def _map_target_market(raw_market: str) -> str | None:
    n = _normalize_market_text(raw_market)
    for canonical, aliases in TARGET_ALIASES.items():
        if any(_normalize_market_text(alias) in n for alias in aliases):
            return canonical
    return None


def _load_cot_file(cot_file: Path) -> pd.DataFrame:
    data = pd.read_csv(cot_file, low_memory=False)
    data = data.dropna(how="all")
    data = data.loc[:, ~data.columns.astype(str).str.startswith("Unnamed")]
    data.columns = [str(col).strip() for col in data.columns]

    market_col = _find_column(data, "market_and_exchange_names")
    date_col = _find_column(data, "report_date_as_yyyy_mm_dd")
    long_col = _find_column(data, "m_money_positions_long_other")
    short_col = _find_column(data, "m_money_positions_short_other")

    missing = []
    if market_col is None:
        missing.append("market_and_exchange_names")
    if date_col is None:
        missing.append("report_date_as_yyyy_mm_dd")
    if long_col is None:
        missing.append("m_money_positions_long_other")
    if short_col is None:
        missing.append("m_money_positions_short_other")
    if missing:
        raise ValueError(f"COT file {cot_file} missing required columns: {missing}")

    cleaned = pd.DataFrame()
    cleaned["market"] = data[market_col].astype(str).str.strip()
    cleaned["cot_report_date"] = pd.to_datetime(data[date_col], errors="coerce", dayfirst=True).dt.normalize()
    cleaned["managed_money_long"] = pd.to_numeric(data[long_col], errors="coerce")
    cleaned["managed_money_short"] = pd.to_numeric(data[short_col], errors="coerce")

    cleaned = cleaned[
        cleaned["market"].ne("")
        & cleaned["cot_report_date"].notna()
        & cleaned["managed_money_long"].notna()
        & cleaned["managed_money_short"].notna()
    ].copy()

    cleaned["market"] = cleaned["market"].apply(_map_target_market)
    cleaned = cleaned[cleaned["market"].notna()].copy()
    cleaned = cleaned.sort_values(["market", "cot_report_date"]).reset_index(drop=True)

    cleaned["managed_money_net"] = cleaned["managed_money_long"] - cleaned["managed_money_short"]
    cleaned["noncommercial_net"] = cleaned["managed_money_net"]
    cleaned["commercial_net"] = cleaned["managed_money_net"]

    grouped = cleaned.groupby("market", sort=False)
    cleaned["weekly_change"] = grouped["managed_money_net"].diff(1)
    cleaned["four_week_change"] = grouped["managed_money_net"].diff(4)
    cleaned["mm_weekly_change"] = grouped["managed_money_net"].diff(1)

    scored = _calculate_cot_scores(cleaned)
    scored["cot_strength"] = scored["cot_strength"].apply(_clean_strength)
    scored["cot_bias"] = scored["cot_bias"].apply(_clean_bias)

    return scored[
        [
            "market",
            "cot_report_date",
            "cot_bias",
            "cot_score",
            "cot_strength",
            "managed_money_net",
            "weekly_change",
            "four_week_change",
        ]
    ]


def _load_cot_history(cot_files: list[Path]) -> pd.DataFrame:
    frames = []
    for f in cot_files:
        print(f"Loading COT file: {f}")
        frames.append(_load_cot_file(f).assign(_cot_source=str(f)))

    history = pd.concat(frames, ignore_index=True)
    history = history.sort_values(["market", "cot_report_date"]).drop_duplicates(
        subset=["market", "cot_report_date"], keep="last"
    )

    if history.empty:
        raise ValueError("COT history is empty after parsing/deduping.")

    
    return history


def _load_macro_history(macro_files: list[Path], sheet_name: str) -> pd.DataFrame:
    frames = []

    for f in macro_files:
        print(f"Loading macro file: {f}")

        try:
            macro = pd.read_excel(f, sheet_name=sheet_name)
        except ValueError:
            print(f"  skipped: {sheet_name} sheet not found")
            continue

        if macro.empty:
            print("  skipped: empty Macro_Dashboard")
            continue

        macro = macro.copy()
        macro.columns = [str(c).strip() for c in macro.columns]

        if "macro_snapshot_date" not in macro.columns:
            print(f"  skipped: missing macro_snapshot_date. Columns: {macro.columns.tolist()}")
            continue

        macro["macro_snapshot_date"] = pd.to_datetime(
            macro["macro_snapshot_date"], errors="coerce"
        ).dt.normalize()

        if "macro_signal" not in macro.columns:
            macro["macro_signal"] = "neutral"
        else:
            macro["macro_signal"] = macro["macro_signal"].astype(str).str.strip().str.lower()

        if "macro_score" not in macro.columns:
            macro["macro_score"] = 0
        else:
            macro["macro_score"] = pd.to_numeric(
                macro["macro_score"], errors="coerce"
            ).fillna(0).clip(0, 10)

        if "macro_strength" not in macro.columns:
            macro["macro_strength"] = "Unknown"

        if "macro_context_for_trades" not in macro.columns:
            macro["macro_context_for_trades"] = ""

        macro["_macro_source"] = str(f)

        usable = macro[
            [
                "macro_snapshot_date",
                "macro_signal",
                "macro_score",
                "macro_strength",
                "macro_context_for_trades",
                "_macro_source",
            ]
        ].copy()

        usable = usable[usable["macro_snapshot_date"].notna()]
        if usable.empty:
            print("  skipped: no usable macro_snapshot_date rows")
            continue

        frames.append(usable)

    if not frames:
        raise ValueError("No usable Macro_Dashboard rows found in macro history files.")

    history = pd.concat(frames, ignore_index=True)
    history = history.sort_values(["macro_snapshot_date", "_macro_source"]).drop_duplicates(
        subset=["macro_snapshot_date"], keep="last"
    )

    if history.empty:
        raise ValueError("Macro history is empty after parsing/deduping.")

    return history


def _build_confluence(cot_bias: str, cot_score: float, macro_signal: str, macro_score: float) -> dict[str, Any]:
    cot_dir = "long" if cot_bias == "Bullish" else "short" if cot_bias == "Bearish" else "neutral"
    macro_dir = "long" if macro_signal == "risk_on" else "short" if macro_signal == "risk_off" else "neutral"
    macro_label = _macro_signal_label(macro_signal)

    hard_conflict = (
        cot_dir in {"long", "short"}
        and macro_dir in {"long", "short"}
        and cot_dir != macro_dir
        and cot_score >= 6
        and macro_score >= 6
    )

    if hard_conflict:
        return {
            "cot_macro_alignment": "Hard Conflict",
            "macro_effect_on_cot": "Blocking",
            "combined_context_score": 0,
            "combined_context_label": "Conflicted / Stand Down",
            "confluence_bias": "Conflicted",
            "confluence_score": 0,
            "confluence_strength": "Blocked",
            "trade_readiness": "Stand down",
            "confluence_read": (
                f"Managed money is {cot_bias.lower()} but macro is {macro_label}, creating a hard conflict with "
                "high-conviction context on both sides."
            ),
            "summary": f"COT {cot_bias} ({cot_score}) conflicts with macro {macro_signal} ({macro_score}) at high conviction.",
        }

    score = cot_score

    directional_macro = cot_dir in {"long", "short"} and macro_dir in {"long", "short"}
    if directional_macro:
        delta = _macro_alignment_adjustment(macro_score)
        score = score + delta if cot_dir == macro_dir else score - delta

    score = max(0, min(10, score))

    if cot_dir == "neutral":
        bias = "Neutral / Mixed"
    elif cot_dir == macro_dir == "long":
        bias = "Long Bias"
    elif cot_dir == macro_dir == "short":
        bias = "Short Bias"
    elif cot_dir == "long":
        bias = "Long (Headwind)"
    else:
        bias = "Short (Headwind)"

    if score >= 8:
        strength = "Very Strong"
        readiness = "High conviction"
    elif score >= 6:
        strength = "Strong"
        readiness = "Actionable"
    elif score >= 3:
        strength = "Moderate"
        readiness = "Cautious"
    else:
        strength = "Weak"
        readiness = "Stand down"

    if cot_dir == "neutral" or macro_dir == "neutral":
        cot_macro_alignment = "Neutral / Mixed"
        macro_effect_on_cot = "Neutral"
    elif cot_dir == macro_dir:
        cot_macro_alignment = "Aligned"
        macro_effect_on_cot = "Boosting"
    else:
        cot_macro_alignment = "Headwind"
        macro_effect_on_cot = "Reducing"

    if score >= 9:
        combined_context_label = "Very Strong Bullish Context" if cot_dir == "long" else "Very Strong Bearish Context"
    elif score >= 7:
        combined_context_label = "Strong Bullish Context" if cot_dir == "long" else "Strong Bearish Context"
    elif score >= 5:
        combined_context_label = "Moderate Bullish Context" if cot_dir == "long" else "Moderate Bearish Context"
    else:
        combined_context_label = "Weak / Neutral Context"

    if cot_dir == "neutral":
        confluence_read = f"Managed money positioning is neutral/mixed and macro is {macro_label}, leaving unclear directional context."
    elif cot_macro_alignment == "Aligned":
        confluence_read = f"Managed money is {cot_bias.lower()} and macro is {macro_label}, giving supportive {cot_dir}-side context."
    elif cot_macro_alignment == "Headwind":
        confluence_read = f"Managed money is {cot_bias.lower()} but macro is {macro_label}, so {cot_dir}-side context faces macro headwind."
    else:
        confluence_read = f"Managed money is {cot_bias.lower()} while macro is {macro_label}, resulting in mixed context."

    return {
        "cot_macro_alignment": cot_macro_alignment,
        "macro_effect_on_cot": macro_effect_on_cot,
        "combined_context_score": score,
        "combined_context_label": combined_context_label,
        "confluence_bias": bias,
        "confluence_score": score,
        "confluence_strength": strength,
        "trade_readiness": readiness,
        "confluence_read": confluence_read,
        "summary": f"COT {cot_bias} ({cot_score}) vs macro {macro_signal} ({macro_score}) => {bias} {score:.1f}.",
    }




def _build_market_diagnostics(cot: pd.DataFrame, out: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    by_market = (
        cot.groupby("market", as_index=False)
        .agg(
            total_rows=("cot_report_date", "size"),
            first_report_date=("cot_report_date", "min"),
            last_report_date=("cot_report_date", "max"),
            latest_available_report_date=("cot_report_date", "max"),
        )
        .sort_values("market")
    )

    date_counts = (
        out.groupby("cot_report_date", as_index=False)
        .agg(rows_on_date=("market", "size"), markets_present=("market", lambda s: ", ".join(sorted(set(s)))))
        .sort_values("cot_report_date")
    )

    all_markets = sorted(set(cot["market"].dropna().astype(str)))
    missing_rows = []
    for week in sorted(out["cot_report_date"].dropna().dt.date.unique()):
        present = set(out[out["cot_report_date"].dt.date == week]["market"].dropna().astype(str))
        for market in all_markets:
            if market not in present:
                reason = "No COT record for this week"
                missing_rows.append({"cot_report_date": week, "missing_market": market, "reason_if_known": reason})
    missing_by_date = pd.DataFrame(missing_rows, columns=["cot_report_date", "missing_market", "reason_if_known"])
    return by_market, date_counts.merge(missing_by_date.groupby("cot_report_date", as_index=False).agg(missing_markets=("missing_market", lambda s: ", ".join(sorted(set(s))))), on="cot_report_date", how="left")

def run() -> Path:
    print("=" * 70)
    print("Confluence history build started")
    print("=" * 70)

    _ensure_cot_backfill(start_year=2023)
    cot_files = _discover_cot_files()
    macro_files, macro_sheet, macro_pattern = _discover_macro_files()

    cot = _load_cot_history(cot_files)
    macro = _load_macro_history(macro_files, macro_sheet)

    print(f"COT date range: {cot['cot_report_date'].min()} -> {cot['cot_report_date'].max()}")
    print(f"Macro date range: {macro['macro_snapshot_date'].min()} -> {macro['macro_snapshot_date'].max()}")
    print("Alignment mode: backward (as-of historical snapshot)")

    aligned = pd.merge_asof(
        cot.sort_values("cot_report_date"),
        macro.sort_values("macro_snapshot_date"),
        left_on="cot_report_date",
        right_on="macro_snapshot_date",
        direction="backward",
    )

    aligned_rows = len(aligned)
    missing_macro_rows = int(aligned["macro_snapshot_date"].isna().sum())
    if missing_macro_rows > 0:
        print("WARNING: Macro history does not fully cover COT history.")
        print(f"  COT date range: {cot['cot_report_date'].min()} -> {cot['cot_report_date'].max()}")
        print(f"  Macro date range: {macro['macro_snapshot_date'].min()} -> {macro['macro_snapshot_date'].max()}")
        print(f"  Rows aligned: {aligned_rows}")
        print(f"  Rows missing macro snapshot: {missing_macro_rows}")

    if aligned.empty:
        raise ValueError("No COT rows could be aligned to available macro snapshots.")

    aligned["macro_signal"] = aligned["macro_signal"].fillna("neutral")
    aligned["macro_score"] = pd.to_numeric(aligned["macro_score"], errors="coerce").fillna(0).clip(0, 10)
    aligned["macro_strength"] = aligned["macro_strength"].fillna("Unknown")
    aligned["macro_context_for_trades"] = aligned["macro_context_for_trades"].fillna("N/A")

    aligned["macro_alignment_gap_days"] = (
        aligned["cot_report_date"] - aligned["macro_snapshot_date"]
    ).dt.days

    confluence_bits = aligned.apply(
        lambda r: _build_confluence(
            cot_bias=str(r["cot_bias"]),
            cot_score=float(r["cot_score"]),
            macro_signal=str(r["macro_signal"]),
            macro_score=float(r["macro_score"]),
        ),
        axis=1,
        result_type="expand",
    )

    out = pd.concat([aligned, confluence_bits], axis=1)
    out = out.sort_values(["market", "cot_report_date"]).reset_index(drop=True)
    out = out[(out["cot_report_date"] >= HISTORY_START_DATE) & (out["cot_report_date"] <= HISTORY_END_DATE)].copy()
    out["confluence_strength"] = out["confluence_score"].apply(_strength_from_score)
    required_columns = [
        "market",
        "cot_report_date",
        "cot_bias",
        "cot_score",
        "cot_strength",
        "macro_snapshot_date",
        "macro_signal",
        "macro_score",
        "macro_strength",
        "macro_context_for_trades",
        "confluence_bias",
        "confluence_score",
        "confluence_strength",
        "trade_readiness",
        "summary",
    ]
    out = out[required_columns + ["_cot_source", "_macro_source", "macro_alignment_gap_days"]].sort_values(
        ["cot_report_date", "market"]
    ).reset_index(drop=True)

    market_diagnostics, date_diagnostics = _build_market_diagnostics(cot, out)

    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = EXPORT_DIR / f"confluence_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    latest_date = out["cot_report_date"].max()
    week_dates = sorted(out["cot_report_date"].dropna().dt.date.unique())
    missing_targets_rows: list[dict[str, Any]] = []
    dashboard = out[required_columns].copy()

    date_index = (
        out.groupby("cot_report_date", as_index=False)
        .agg(
            markets_present=("market", lambda s: ", ".join(sorted(set(s)))),
            avg_confluence_score=("confluence_score", "mean"),
            actionable_count=("trade_readiness", lambda s: int((s == "Actionable").sum())),
            high_conviction_count=("trade_readiness", lambda s: int((s == "High conviction").sum())),
        )
        .sort_values("cot_report_date")
    )
    strongest_bull = (
        out[out["confluence_bias"].isin(["Long Bias", "Long (Headwind)"])]
        .sort_values(["cot_report_date", "confluence_score"], ascending=[True, False])
        .drop_duplicates("cot_report_date")[["cot_report_date", "market", "confluence_score"]]
        .rename(columns={"market": "strongest_bullish_market"})
    )
    strongest_bear = (
        out[out["confluence_bias"].isin(["Short Bias", "Short (Headwind)"])]
        .sort_values(["cot_report_date", "confluence_score"], ascending=[True, False])
        .drop_duplicates("cot_report_date")[["cot_report_date", "market", "confluence_score"]]
        .rename(columns={"market": "strongest_bearish_market"})
    )
    date_index = date_index.merge(strongest_bull, on="cot_report_date", how="left").merge(
        strongest_bear, on="cot_report_date", how="left"
    )
    date_index["avg_confluence_score"] = date_index["avg_confluence_score"].round(2)

    for week in week_dates:
        present = set(out[out["cot_report_date"].dt.date == week]["market"].dropna().astype(str))
        missing = sorted(set(TARGET_MARKETS) - present)
        for market in missing:
            missing_targets_rows.append(
                {"missing_target_market": market, "aliases_searched": market, "reason_if_known": "No COT record for this week"}
            )
    missing_targets = pd.DataFrame(
        missing_targets_rows, columns=["missing_target_market", "aliases_searched", "reason_if_known"]
    ).drop_duplicates()

    source_notes = pd.DataFrame(
        [
            ("cot_files_used", "; ".join(str(f) for f in cot_files)),
            ("macro_files_used", "; ".join(str(f) for f in macro_files)),
            ("macro_reconstruction_date_range", f"{macro['macro_snapshot_date'].min()} -> {macro['macro_snapshot_date'].max()}"),
            ("confluence_date_range", f"{out['cot_report_date'].min()} -> {out['cot_report_date'].max()}"),
            ("row_count", len(out)),
            ("known_limitations", "Rows emitted for every COT market/date; missing macro snapshots are retained with neutral/0 defaults and null snapshot date."),
        ],
        columns=["field", "value"],
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        dashboard.to_excel(writer, sheet_name="Confluence_Dashboard", index=False)
        date_index.to_excel(writer, sheet_name="Date_Index", index=False)
        out[required_columns].to_excel(writer, sheet_name="Weekly_Blocks", index=False)
        missing_targets.to_excel(writer, sheet_name="Missing_Targets", index=False)
        source_notes.to_excel(writer, sheet_name="Source_Notes", index=False)
        market_diagnostics.to_excel(writer, sheet_name="Market_Diagnostics", index=False)
        date_diagnostics.to_excel(writer, sheet_name="Date_Diagnostics", index=False)

    wb = load_workbook(output_path)
    dashboard_ws = wb["Confluence_Dashboard"]
    date_index_ws = wb["Date_Index"]
    weekly_ws = wb["Weekly_Blocks"]
    missing_targets_ws = wb["Missing_Targets"]
    source_notes_ws = wb["Source_Notes"]

    def _format_table_sheet(ws) -> None:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for col_cells in ws.columns:
            width = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(width + 2, 80)

    for ws in [dashboard_ws, date_index_ws, weekly_ws, missing_targets_ws, source_notes_ws]:
        _format_table_sheet(ws)
    weekly_ws.delete_rows(1, weekly_ws.max_row)
    row_ptr = 1
    for week in sorted(out["cot_report_date"].dropna().dt.date.unique()):
        weekly_ws.cell(row=row_ptr, column=1, value=f"Week of {week}").font = Font(bold=True, size=14)
        row_ptr += 1
        for i, col in enumerate(required_columns, start=1):
            weekly_ws.cell(row=row_ptr, column=i, value=col).font = Font(bold=True)
        row_ptr += 1
        week_rows = out[out["cot_report_date"].dt.date == week][required_columns].sort_values("market")
        for _, r in week_rows.iterrows():
            for i, col in enumerate(required_columns, start=1):
                weekly_ws.cell(row=row_ptr, column=i, value=r[col])
            row_ptr += 1
        row_ptr += 1

    green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    amber_fill = PatternFill(fill_type="solid", fgColor="FFEB9C")
    red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")

    def _apply_confluence_colors(ws) -> None:
        header_to_index = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
        score_col = header_to_index.get("confluence_score")
        readiness_col = header_to_index.get("trade_readiness")
        bias_col = header_to_index.get("confluence_bias")
        if score_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=score_col)
                value = cell.value
                if isinstance(value, (int, float)):
                    if value >= 8:
                        cell.fill = green_fill
                    elif value >= 5:
                        cell.fill = amber_fill
                    else:
                        cell.fill = red_fill
        if readiness_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=readiness_col)
                value = str(cell.value).strip()
                if value == "High conviction":
                    cell.fill = green_fill
                elif value in {"Actionable", "Cautious"}:
                    cell.fill = amber_fill
                elif value == "Stand down":
                    cell.fill = red_fill
        if bias_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=bias_col)
                value = str(cell.value).strip()
                if value in {"Long Bias", "Short Bias"}:
                    cell.fill = green_fill
                elif value in {"Long (Headwind)", "Short (Headwind)"}:
                    cell.fill = amber_fill
                elif value == "Conflicted":
                    cell.fill = red_fill

    _apply_confluence_colors(dashboard_ws)
    _apply_confluence_colors(weekly_ws)
    dashboard_ws.freeze_panes = "A2"
    latest_rows = out[out["cot_report_date"] == latest_date][["market", "confluence_score"]].sort_values("market")
    charts_sheet = wb.create_sheet("Summary_Charts")
    charts_sheet["A1"] = f"Latest Week Confluence ({latest_date.date()})"
    charts_sheet["A1"].font = Font(bold=True, size=13)
    charts_sheet["A3"] = "market"
    charts_sheet["B3"] = "confluence_score"
    for idx, (_, r) in enumerate(latest_rows.iterrows(), start=4):
        charts_sheet.cell(row=idx, column=1, value=r["market"])
        charts_sheet.cell(row=idx, column=2, value=float(r["confluence_score"]))
    chart = BarChart()
    chart.title = f"Confluence Scores - {latest_date.date()}"
    chart.y_axis.title = "Confluence Score"
    chart.x_axis.title = "Market"
    data_ref = Reference(charts_sheet, min_col=2, min_row=3, max_row=3 + len(latest_rows))
    cat_ref = Reference(charts_sheet, min_col=1, min_row=4, max_row=3 + len(latest_rows))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    chart.height = 9
    chart.width = 18
    charts_sheet.add_chart(chart, "D3")

    wb.save(output_path)

    print("=" * 70)
    print(f"COT input files used: {len(cot_files)}")
    for f in cot_files:
        print(f"  - {f}")
    print(f"Macro input files used ({macro_pattern}, sheet={macro_sheet}): {len(macro_files)}")
    for f in macro_files:
        print(f"  - {f}")
    confluence_weeks = out["cot_report_date"].nunique()
    confluence_markets = out["market"].nunique()
    present_markets = set(out["market"].dropna().unique())
    missing_target_markets = sorted(set(TARGET_MARKETS) - present_markets)

    print(f"Confluence date range: {out['cot_report_date'].min()} -> {out['cot_report_date'].max()}")
    print(f"Weeks generated: {confluence_weeks}")
    print(f"Instruments found: {confluence_markets}")
    print(f"Aliases searched: {TARGET_ALIASES}")
    print(f"Aliases matched: {sorted(out['market'].dropna().unique())}")
    print(f"Rows exported: {len(out)}")
    if missing_target_markets:
        print("Instruments missing:")
        for market in missing_target_markets:
            print(f"  - {market}")
    else:
        print("Instruments missing: none")
    print(f"Output file path: {output_path}")
    print("=" * 70)

    return output_path


if __name__ == "__main__":
    run()
