from __future__ import annotations

from datetime import datetime
import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from hptl.history.export_confluence_json import run as export_dashboard_json

EXPORT_DIR = Path("data/exports")
PROCESSED_DIR = Path("data/processed")


def _latest_file(paths: list[Path]) -> Path:
    existing = [p for p in paths if p.exists()]
    if not existing:
        raise FileNotFoundError("No candidate files found.")
    return max(existing, key=lambda p: p.stat().st_mtime)


def _latest_cot_file() -> Path:
    candidates = list(EXPORT_DIR.glob("cot_update_*.xlsx")) + list(
        PROCESSED_DIR.glob("cot_cleaned_*.csv")
    )
    if not candidates:
        raise FileNotFoundError(
            "No COT input found. Expected data/exports/cot_update_*.xlsx or data/processed/cot_cleaned_*.csv"
        )
    return max(candidates, key=lambda p: p.stat().st_mtime)


def _normalize_column_name(col: str) -> str:
    return " ".join(str(col).strip().lower().replace("_", " ").split())


def _find_column(df: pd.DataFrame, *aliases: str) -> str | None:
    normalized = {_normalize_column_name(col): col for col in df.columns}
    for alias in aliases:
        col = normalized.get(_normalize_column_name(alias))
        if col is not None:
            return col
    return None


def _clean_bias(value: Any) -> str:
    text = str(value).strip().lower()
    if "bull" in text or text in {"long", "buy"}:
        return "Bullish"
    if "bear" in text or text in {"short", "sell"}:
        return "Bearish"
    return "Neutral / Mixed"


def _clean_strength(value: Any) -> str:
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return "Unknown"
    return text


def _strength_to_score(strength: Any) -> float:
    text = str(strength).strip().lower()
    if "very strong" in text or "strongly" in text:
        return 8
    if text == "strong":
        return 7
    if text == "moderate":
        return 5
    if text == "weak":
        return 2
    return 0


def _load_cot(cot_file: Path) -> pd.DataFrame:
    if cot_file.suffix.lower() == ".xlsx":
        raw = pd.read_excel(cot_file, sheet_name="Dashboard", header=None)
        header_idx = None
        for idx, row in raw.iterrows():
            if row.astype(str).str.strip().eq("Market").any():
                header_idx = idx
                break
        if header_idx is None:
            raise ValueError("Could not find 'Market' header row in COT Dashboard sheet.")

        header = raw.iloc[header_idx].astype(str).str.strip().tolist()
        data = raw.iloc[header_idx + 1 :].copy()
        data.columns = header
    else:
        data = pd.read_csv(cot_file)

    data = data.dropna(how="all")
    data = data.loc[:, ~data.columns.astype(str).str.startswith("Unnamed")]
    data.columns = [str(col).strip() for col in data.columns]
    print(f"COT columns found: {list(data.columns)}")

    market_col = _find_column(
        data,
        "market",
        "market_name",
        "instrument",
        "symbol",
        "contract_market_name",
        "market_and_exchange_names",
    )
    date_col = _find_column(data, "cot_report_date", "report date", "date")
    bias_col = _find_column(data, "cot_bias", "bias", "signal")
    score_col = _find_column(data, "cot_score", "score")
    strength_col = _find_column(data, "cot_strength", "strength")

    if market_col is None:
        raise ValueError(
            "COT data is missing a market/instrument column. "
            "Expected one of: Market, market, market_name, Market_Name, instrument, symbol, "
            "contract_market_name, market_and_exchange_names. "
            f"Available columns: {list(data.columns)}"
        )
    print(f"Market column detected: {market_col}")

    cleaned = pd.DataFrame()
    cleaned["market"] = data[market_col].astype(str).str.strip()
    cleaned = cleaned[cleaned["market"].ne("")]

    if date_col is not None:
        cleaned["cot_report_date"] = pd.to_datetime(data[date_col], errors="coerce").dt.date
    else:
        cleaned["cot_report_date"] = pd.NaT

    cleaned["cot_bias"] = data[bias_col].apply(_clean_bias) if bias_col else pd.NA

    if score_col:
        cleaned["cot_score"] = pd.to_numeric(data[score_col], errors="coerce")
    else:
        cleaned["cot_score"] = pd.NA

    cleaned["cot_strength"] = data[strength_col].apply(_clean_strength) if strength_col else "Unknown"

    missing_score = cleaned["cot_score"].isna()
    fallback_scores = pd.to_numeric(
        cleaned.loc[missing_score, "cot_strength"].apply(_strength_to_score),
        errors="coerce",
    )
    cleaned.loc[missing_score, "cot_score"] = fallback_scores
    cleaned["cot_score"] = pd.to_numeric(cleaned["cot_score"], errors="coerce").clip(lower=0, upper=10)

    cleaned = cleaned[cleaned["cot_bias"].notna() & cleaned["cot_score"].notna()].copy()

    return cleaned.reset_index(drop=True)


def _macro_alignment_adjustment(score: float) -> int:
    if score <= 2:
        return 1
    if score <= 5:
        return 2
    if score <= 7:
        return 3
    return 4


def _to_num(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def _write_audits(cot: pd.DataFrame, confluence: pd.DataFrame, dashboard_json_path: Path) -> None:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    cot2 = cot.copy()
    cot2["cot_report_date"] = pd.to_datetime(cot2["cot_report_date"], errors="coerce")

    coverage = (
        cot2.groupby("market", dropna=False)
        .agg(
            first_report_date=("cot_report_date", "min"),
            last_report_date=("cot_report_date", "max"),
            row_count=("market", "size"),
        )
        .reset_index()
    )
    dashboard_markets = set(confluence["market"].astype(str).str.strip().tolist())
    coverage["included_in_dashboard"] = coverage["market"].astype(str).str.strip().isin(dashboard_markets)
    coverage["missing_reason"] = coverage["included_in_dashboard"].map(
        lambda included: "" if included else "not present in confluence output"
    )
    coverage["first_report_date"] = pd.to_datetime(coverage["first_report_date"], errors="coerce").dt.date
    coverage["last_report_date"] = pd.to_datetime(coverage["last_report_date"], errors="coerce").dt.date
    coverage.to_csv(EXPORT_DIR / "market_coverage_audit.csv", index=False)

    cocoa = confluence[confluence["market"].astype(str).str.contains("COCOA", case=False, na=False)].copy()
    cocoa = cocoa.sort_values("cot_report_date")
    cocoa["report_date"] = pd.to_datetime(cocoa.get("cot_report_date"), errors="coerce").dt.date
    cocoa["managed_money_long"] = _to_num(cocoa.get("noncommercial_long", pd.Series(index=cocoa.index)))
    cocoa["managed_money_short"] = _to_num(cocoa.get("noncommercial_short", pd.Series(index=cocoa.index)))
    cocoa["managed_money_net"] = _to_num(cocoa.get("noncommercial_net", pd.Series(index=cocoa.index)))
    cocoa["1w_change"] = _to_num(cocoa.get("mm_weekly_change", pd.Series(index=cocoa.index)))
    cocoa["4w_change"] = _to_num(cocoa.get("mm_four_week_change", pd.Series(index=cocoa.index)))
    cocoa["final_bias"] = cocoa.get("confluence_bias")
    cocoa["final_score"] = _to_num(cocoa.get("confluence_score", pd.Series(index=cocoa.index)))
    cocoa["dashboard_bias"] = cocoa.get("confluence_bias")
    cocoa["dashboard_score"] = cocoa.get("confluence_score")

    cols = [
        "report_date",
        "market",
        "noncommercial_long",
        "noncommercial_short",
        "noncommercial_net",
        "managed_money_long",
        "managed_money_short",
        "managed_money_net",
        "1w_change",
        "4w_change",
        "cot_bias",
        "cot_score",
        "macro_signal",
        "macro_score",
        "final_bias",
        "final_score",
        "dashboard_bias",
        "dashboard_score",
    ]
    for column in cols:
        if column not in cocoa.columns:
            cocoa[column] = pd.NA
    cocoa[cols].to_csv(EXPORT_DIR / "cocoa_score_audit.csv", index=False)

    json_records = json.loads(dashboard_json_path.read_text(encoding="utf-8")).get("records", [])
    print(f"dashboard JSON row count: {len(json_records)}")
    cot_markets = sorted(cot2["market"].dropna().astype(str).str.strip().unique().tolist())
    dash_markets = sorted(confluence["market"].dropna().astype(str).str.strip().unique().tolist())
    print(f"all markets found in COT input: {cot_markets}")
    print(f"all markets written to dashboard JSON: {dash_markets}")

    latest_cocoa = cocoa.dropna(subset=["report_date"]).tail(1)
    if latest_cocoa.empty:
        print("Cocoa latest COT score and bias: not found")
    else:
        row = latest_cocoa.iloc[0]
        print(f"Cocoa latest COT score and bias: {row.get('cot_score')} / {row.get('cot_bias')}")
        if (
            pd.notna(row.get("final_score"))
            and pd.notna(row.get("cot_score"))
            and float(row.get("final_score")) != float(row.get("cot_score"))
        ):
            print(
                "reason Cocoa final score differs from raw COT score, if it does: "
                f"macro confluence adjustment using macro_signal={row.get('macro_signal')} "
                f"and macro_score={row.get('macro_score')}"
            )
        else:
            print("reason Cocoa final score differs from raw COT score, if it does: no difference (or missing values)")


def run() -> None:
    print("=" * 70)
    print("Confluence update started")
    print("=" * 70)

    cot_file = _latest_cot_file()
    macro_file = _latest_file(list(EXPORT_DIR.glob("macro_output_*.xlsx")))

    print(f"COT file used: {cot_file}")
    print(f"Macro file used: {macro_file}")

    cot = _load_cot(cot_file)
    print(f"Cleaned COT rows: {len(cot)}")
    rows_by_market = cot["market"].value_counts(dropna=False).sort_index()
    print("rows by market:")
    print(rows_by_market.to_string())
    latest_report_date = pd.to_datetime(cot["cot_report_date"], errors="coerce").max()
    if pd.isna(latest_report_date):
        missing_markets: list[str] = []
    else:
        all_markets = set(
            cot["market"].astype(str).str.strip().replace("", pd.NA).dropna().tolist()
        )
        latest_mask = pd.to_datetime(cot["cot_report_date"], errors="coerce") == latest_report_date
        latest_markets = set(
            cot.loc[latest_mask, "market"]
            .astype(str)
            .str.strip()
            .replace("", pd.NA)
            .dropna()
            .tolist()
        )
        missing_markets = sorted(all_markets - latest_markets)
    latest_label = "N/A" if pd.isna(latest_report_date) else str(latest_report_date.date())
    print(f"missing markets by latest report date ({latest_label}): {missing_markets}")

    macro = pd.read_excel(macro_file, sheet_name="Macro_Dashboard")
    if macro.empty:
        raise ValueError("Macro_Dashboard sheet is empty.")

    macro_row = macro.iloc[-1]
    macro_score = float(pd.to_numeric(macro_row.get("macro_score"), errors="coerce") or 0)
    macro_signal = str(macro_row.get("macro_signal", "")).strip().lower()

    print(f"Macro signal / score used: {macro_signal} / {macro_score}")

    macro_snapshot_date = pd.to_datetime(macro_row.get("snapshot_date"), errors="coerce")
    if pd.isna(macro_snapshot_date):
        macro_snapshot_date = pd.Timestamp(datetime.utcnow().date())

    def build_confluence(row: pd.Series) -> pd.Series:
        cot_bias = row["cot_bias"]
        cot_score = float(row["cot_score"])

        cot_dir = "long" if cot_bias == "Bullish" else "short" if cot_bias == "Bearish" else "neutral"
        macro_dir = "long" if macro_signal == "risk_on" else "short" if macro_signal == "risk_off" else "neutral"

        hard_conflict = (
            cot_dir in {"long", "short"}
            and macro_dir in {"long", "short"}
            and cot_dir != macro_dir
            and cot_score >= 6
            and macro_score >= 6
        )

        if hard_conflict:
            return pd.Series(
                {
                    "confluence_bias": "Conflicted / No Trade",
                    "confluence_score": 0,
                    "confluence_strength": "Blocked",
                    "trade_readiness": "Stand down",
                    "summary": f"{cot_bias} COT conflicts with {macro_signal} macro at high conviction.",
                }
            )

        score = cot_score
        if cot_dir in {"long", "short"} and macro_dir in {"long", "short"}:
            delta = _macro_alignment_adjustment(macro_score)
            score = score + delta if cot_dir == macro_dir else score - delta

        score = max(0, min(10, score))

        if cot_dir == "neutral":
            bias = "Neutral / Mixed"
        elif cot_dir == macro_dir == "long":
            bias = "Long Bias"
        elif cot_dir == macro_dir == "short":
            bias = "Short Bias"
        elif cot_dir == "long":
            bias = "Long (Headwind)"
        else:
            bias = "Short (Headwind)"

        if score >= 8:
            strength = "Very Strong"
            readiness = "High conviction"
        elif score >= 6:
            strength = "Strong"
            readiness = "Actionable"
        elif score >= 3:
            strength = "Moderate"
            readiness = "Cautious"
        else:
            strength = "Weak"
            readiness = "Low conviction"

        return pd.Series(
            {
                "confluence_bias": bias,
                "confluence_score": score,
                "confluence_strength": strength,
                "trade_readiness": readiness,
                "summary": f"COT {cot_bias} ({cot_score}) vs macro {macro_signal} ({macro_score}) => {bias} {score:.1f}.",
            }
        )

    confluence = cot.copy()
    confluence["macro_snapshot_date"] = macro_snapshot_date.date()
    confluence["macro_signal"] = macro_signal
    confluence["macro_score"] = macro_score
    confluence["macro_strength"] = macro_row.get("macro_strength")
    confluence["macro_context_for_trades"] = macro_row.get("macro_context_for_trades")

    confluence = pd.concat([confluence, confluence.apply(build_confluence, axis=1)], axis=1)

    required_output_columns = [
        "confluence_bias",
        "confluence_score",
        "cot_bias",
        "cot_score",
        "macro_signal",
        "macro_score",
        "trade_readiness",
        "summary",
    ]

    fallback_map = {
        "confluence_bias": ("final_bias", "bias", "direction"),
        "confluence_score": ("final_score", "score"),
        "cot_bias": ("bias", "final_bias", "direction"),
        "cot_score": ("score", "final_score"),
        "macro_signal": ("final_context", "macro_context", "context"),
        "macro_score": ("score", "final_score"),
        "trade_readiness": ("readiness", "status"),
        "summary": ("final_context", "context", "notes"),
    }

    for target_col in required_output_columns:
        if target_col in confluence.columns:
            continue
        for source_col in fallback_map.get(target_col, ()):
            if source_col in confluence.columns:
                confluence[target_col] = confluence[source_col]
                break
        else:
            confluence[target_col] = None

    confluence["confluence_bias"] = confluence["confluence_bias"].fillna("Conflicted / No Trade")

    print("Confluence dataframe columns before sorting:", sorted(confluence.columns.tolist()))

    bias_order = {
        "Long Bias": 0,
        "Short Bias": 0,
        "Long (Headwind)": 1,
        "Short (Headwind)": 1,
        "Neutral / Mixed": 2,
        "Conflicted / No Trade": 3,
    }
    confluence["_bias_rank"] = confluence["confluence_bias"].map(bias_order).fillna(4)
    confluence["confluence_score"] = pd.to_numeric(confluence["confluence_score"], errors="coerce")
    confluence = confluence.sort_values(
        by=["_bias_rank", "confluence_score"], ascending=[True, False]
    ).drop(columns=["_bias_rank"])

    final_columns = [
        "market",
        "cot_report_date",
        "cot_bias",
        "cot_score",
        "cot_strength",
        "macro_snapshot_date",
        "macro_signal",
        "macro_score",
        "macro_strength",
        "macro_context_for_trades",
        "confluence_bias",
        "confluence_score",
        "confluence_strength",
        "trade_readiness",
        "summary",
    ]
    confluence = confluence[final_columns]

    output_path = EXPORT_DIR / f"confluence_output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    source_notes = pd.DataFrame(
        {
            "field": ["cot_file_used", "macro_file_used", "generated_at_utc"],
            "value": [str(cot_file), str(macro_file), datetime.utcnow().isoformat()],
        }
    )

    with pd.ExcelWriter(output_path) as writer:
        confluence.to_excel(writer, sheet_name="Confluence_Dashboard", index=False)
        cot.to_excel(writer, sheet_name="COT_Input", index=False)
        macro.to_excel(writer, sheet_name="Macro_Input", index=False)
        source_notes.to_excel(writer, sheet_name="Source_Notes", index=False)

    wb = load_workbook(output_path)
    dashboard_ws = wb["Confluence_Dashboard"]
    dashboard_ws.freeze_panes = "A2"
    dashboard_ws.auto_filter.ref = dashboard_ws.dimensions

    bold_font = Font(bold=True)
    for cell in dashboard_ws[1]:
        cell.font = bold_font

    for col_cells in dashboard_ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        dashboard_ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 80)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    amber_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    max_row = dashboard_ws.max_row

    trade_col = get_column_letter(confluence.columns.get_loc("trade_readiness") + 1)
    trade_range = f"{trade_col}2:{trade_col}{max_row}"
    dashboard_ws.conditional_formatting.add(trade_range, FormulaRule(formula=[f'ISNUMBER(SEARCH("High conviction",{trade_col}2))'], fill=green_fill))
    dashboard_ws.conditional_formatting.add(trade_range, FormulaRule(formula=[f'ISNUMBER(SEARCH("Actionable",{trade_col}2))'], fill=green_fill))
    dashboard_ws.conditional_formatting.add(trade_range, FormulaRule(formula=[f'ISNUMBER(SEARCH("Cautious",{trade_col}2))'], fill=amber_fill))
    for term in ["Conflicted", "Blocked", "No Trade"]:
        dashboard_ws.conditional_formatting.add(trade_range, FormulaRule(formula=[f'ISNUMBER(SEARCH("{term}",{trade_col}2))'], fill=red_fill))

    score_col = get_column_letter(confluence.columns.get_loc("confluence_score") + 1)
    score_range = f"{score_col}2:{score_col}{max_row}"
    dashboard_ws.conditional_formatting.add(score_range, CellIsRule(operator="between", formula=["8", "10"], fill=green_fill))
    dashboard_ws.conditional_formatting.add(score_range, CellIsRule(operator="between", formula=["5", "7"], fill=amber_fill))
    dashboard_ws.conditional_formatting.add(score_range, CellIsRule(operator="between", formula=["0", "4"], fill=red_fill))

    if "Summary_Charts" in wb.sheetnames:
        del wb["Summary_Charts"]
    summary_ws = wb.create_sheet("Summary_Charts")

    score_summary = confluence[["market", "confluence_score"]].copy()
    readiness_counts = confluence["trade_readiness"].value_counts(dropna=False).rename_axis("trade_readiness").reset_index(name="count")
    bias_counts = confluence["confluence_bias"].value_counts(dropna=False).rename_axis("confluence_bias").reset_index(name="count")

    summary_ws["A1"] = "market"
    summary_ws["B1"] = "confluence_score"
    for i, row in enumerate(score_summary.itertuples(index=False), start=2):
        summary_ws.cell(row=i, column=1, value=row.market)
        summary_ws.cell(row=i, column=2, value=float(row.confluence_score) if pd.notna(row.confluence_score) else None)
    score_chart = BarChart()
    score_chart.title = "Confluence Score by Market"
    score_chart.add_data(Reference(summary_ws, min_col=2, min_row=1, max_row=1 + len(score_summary)), titles_from_data=True)
    score_chart.set_categories(Reference(summary_ws, min_col=1, min_row=2, max_row=1 + len(score_summary)))
    summary_ws.add_chart(score_chart, "D2")

    readiness_start = len(score_summary) + 4
    summary_ws.cell(row=readiness_start, column=1, value="trade_readiness")
    summary_ws.cell(row=readiness_start, column=2, value="count")
    for i, row in enumerate(readiness_counts.itertuples(index=False), start=readiness_start + 1):
        summary_ws.cell(row=i, column=1, value=row.trade_readiness)
        summary_ws.cell(row=i, column=2, value=int(row.count))
    readiness_chart = BarChart()
    readiness_chart.title = "Trade Readiness Categories"
    readiness_chart.add_data(Reference(summary_ws, min_col=2, min_row=readiness_start, max_row=readiness_start + len(readiness_counts)), titles_from_data=True)
    readiness_chart.set_categories(Reference(summary_ws, min_col=1, min_row=readiness_start + 1, max_row=readiness_start + len(readiness_counts)))
    summary_ws.add_chart(readiness_chart, f"D{readiness_start}")

    bias_start = readiness_start + len(readiness_counts) + 3
    summary_ws.cell(row=bias_start, column=1, value="confluence_bias")
    summary_ws.cell(row=bias_start, column=2, value="count")
    for i, row in enumerate(bias_counts.itertuples(index=False), start=bias_start + 1):
        summary_ws.cell(row=i, column=1, value=row.confluence_bias)
        summary_ws.cell(row=i, column=2, value=int(row.count))
    bias_chart = BarChart()
    bias_chart.title = "Confluence Bias Categories"
    bias_chart.add_data(Reference(summary_ws, min_col=2, min_row=bias_start, max_row=bias_start + len(bias_counts)), titles_from_data=True)
    bias_chart.set_categories(Reference(summary_ws, min_col=1, min_row=bias_start + 1, max_row=bias_start + len(bias_counts)))
    summary_ws.add_chart(bias_chart, f"D{bias_start}")

    for col_idx in range(1, 9):
        summary_ws.column_dimensions[get_column_letter(col_idx)].width = 28 if col_idx == 1 else 16
    for cell in summary_ws[1]:
        cell.font = bold_font
    wb.save(output_path)

    print(f"Output path saved: {output_path}")
    dashboard_json_path = export_dashboard_json(input_path=str(output_path))
    _write_audits(cot, confluence, dashboard_json_path)
    print(f"Dashboard JSON refreshed from historical workbook: {dashboard_json_path}")


if __name__ == "__main__":
    run()
