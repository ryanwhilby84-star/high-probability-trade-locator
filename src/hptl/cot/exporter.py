from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from hptl.config import Settings
from hptl.cot.contracts import CME_FUTURES_ONLY_URL, FINANCIAL_FUTURES_ONLY_URL_TEMPLATE, GOOD_WORKBOOK_DISPLAY_NAMES, GOOD_WORKBOOK_MARKET_ORDER, LEGACY_FUTURES_ONLY_URL_TEMPLATE
from hptl.cot.parser import markets_included
from hptl.shared.file_utils import ensure_dir


DASHBOARD_COLUMNS = [
    "report_date",
    "market_name",
    "exchange",
    "open_interest",
    "noncommercial_long",
    "noncommercial_short",
    "commercial_long",
    "commercial_short",
    "commercial_net",
    "noncommercial_net",
    "weekly_change",
    "four_week_change",
    "bias",
]

OPTIONAL_EXPORT_COLUMNS = [
    "dashboard_market",
    "market_name_clean",
    "cftc_contract_market_code",
    "source_report",
    "asset_class",
]

TRADER_REPORT_COLUMNS = [
    "Date",
    "Market",
    "CFTC Contract",
    "Open Interest",
    "Commercial Long",
    "Commercial Short",
    "Commercial Net",
    "Commercial Net 1W Chg",
    "Commercial Net 4W Chg",
    "Managed Money Long",
    "Managed Money Short",
    "Managed Money Net",
    "MM Net 1W Chg",
    "MM Net 4W Chg",
    "Bias",
    "cot_bias",
    "cot_score",
    "cot_strength",
    "cot_summary",
]

MARKET_BLOCK_COLUMNS = [
    "Date",
    "Commercial Long",
    "Commercial Short",
    "Commercial Net",
    "1W Chg",
    "4W Chg",
    "Managed Net",
    "MM 1W Chg",
    "Bias",
    "cot_bias",
    "cot_score",
    "cot_strength",
    "cot_summary",
]

NAVY = "0F172A"
HEADER_NAVY = "111827"
WHITE = "FFFFFF"
GREEN_FILL = "D9EAD3"
RED_FILL = "F4CCCC"
YELLOW_FILL = "FFF2CC"
POS_FILL = "D1FAE5"
NEG_FILL = "FEE2E2"
STRENGTH_WEAK_FILL = "E5E7EB"
STRENGTH_MODERATE_FILL = "FFF2CC"
STRENGTH_STRONG_FILL = "D9EAD3"
STRENGTH_VERY_STRONG_FILL = "38761D"
GRID = "E5E7EB"


@dataclass(frozen=True)
class ExportResult:
    export_file_path: Path
    processed_csv_path: Path
    rows_exported: int
    markets: list[str]


def _normalise_dashboard_input(df: pd.DataFrame) -> pd.DataFrame:
    dashboard = df.copy()
    if "dashboard_market" in dashboard.columns:
        dashboard["market_name"] = dashboard["dashboard_market"].fillna(dashboard.get("market_name"))
    elif "market_name_clean" in dashboard.columns and "market_name" not in dashboard.columns:
        dashboard["market_name"] = dashboard["market_name_clean"]

    for column in DASHBOARD_COLUMNS + OPTIONAL_EXPORT_COLUMNS:
        if column not in dashboard.columns:
            dashboard[column] = pd.NA

    dashboard = dashboard[DASHBOARD_COLUMNS + OPTIONAL_EXPORT_COLUMNS]
    dashboard = dashboard[dashboard["market_name"].notna()].copy()
    dashboard["market_name"] = dashboard["market_name"].astype(str).str.strip()
    dashboard = dashboard[dashboard["market_name"].ne("")]
    return dashboard


def _canonical_market_label(market_name: object) -> str:
    text = str(market_name).strip().upper()
    aliases = {
        "CRUDE OIL (WTI)": "CRUDE OIL",
        "WTI CRUDE OIL": "CRUDE OIL",
        "E-MINI NASDAQ 100": "NASDAQ",
        "NQ": "NASDAQ",
        "ES": "S&P 500",
        "S&P500": "S&P 500",
        "SP 500": "S&P 500",
    }
    return aliases.get(text, text)


def _market_sort_priority(market_name: object) -> int:
    text = _canonical_market_label(market_name)
    try:
        return GOOD_WORKBOOK_MARKET_ORDER.index(text)
    except ValueError:
        return 999


def _sort_workbook_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Sort by the locked workbook market order, then by date ascending.

    This is used before any time-series calculations and before every output
    view. It deliberately avoids relying on the incoming CFTC file order.
    """
    sorted_df = df.copy()
    sorted_df["market_name"] = sorted_df["market_name"].map(_canonical_market_label)
    sorted_df["report_date"] = pd.to_datetime(sorted_df["report_date"], errors="coerce")
    sorted_df["_sort_priority"] = sorted_df["market_name"].apply(_market_sort_priority)
    sorted_df = sorted_df.sort_values(
        ["_sort_priority", "market_name", "report_date"],
        ascending=[True, True, True],
        na_position="last",
        kind="mergesort",
    )
    return sorted_df.drop(columns=["_sort_priority"]).reset_index(drop=True)


def _deduplicate_for_calculation(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """Keep one row per market/date before calculating week changes.

    When the CFTC source provides duplicate aliases for a contract week, keep the
    row with the most required numeric inputs populated. Duplicates are removed
    before diff/shift so one market never gets an artificial zero/extra week.
    """
    if df.empty or not {"market_name", "report_date"}.issubset(df.columns):
        return df.copy(), []

    result = df.copy()
    result["report_date"] = pd.to_datetime(result["report_date"], errors="coerce")
    value_cols = [
        "open_interest",
        "commercial_long",
        "commercial_short",
        "noncommercial_long",
        "noncommercial_short",
    ]
    existing = [col for col in value_cols if col in result.columns]
    result["_populated_values"] = result[existing].notna().sum(axis=1) if existing else 0
    before = len(result)
    result = result.sort_values(
        ["market_name", "report_date", "_populated_values"],
        ascending=[True, True, False],
        kind="mergesort",
    )
    result = result.drop_duplicates(["market_name", "report_date"], keep="first")
    removed = before - len(result)
    warnings = [f"Removed {removed} duplicate market/date rows before calculations; kept the most populated row for each market/date."] if removed else []
    return result.drop(columns=["_populated_values"]).reset_index(drop=True), warnings


def _validate_market_date_order(df: pd.DataFrame) -> list[str]:
    """Validate each market is a strictly increasing time series."""
    warnings: list[str] = []
    if df.empty:
        return warnings

    for market in GOOD_WORKBOOK_MARKET_ORDER:
        dates = pd.to_datetime(df.loc[df["market_name"] == market, "report_date"], errors="coerce").dropna().reset_index(drop=True)
        if len(dates) <= 1:
            continue

        diffs = dates.diff().dropna()
        if not (diffs > pd.Timedelta(0)).all():
            warnings.append(
                f"Date ordering warning for {GOOD_WORKBOOK_DISPLAY_NAMES.get(market, market)}: "
                "report_date values are not strictly increasing after sorting/deduplication."
            )

        # Weekly COT reports should normally move in 7-day steps. This does not
        # fake any values; it simply flags missing source weeks if present.
        gap_dates = dates[dates.diff() > pd.Timedelta(days=10)]
        if not gap_dates.empty:
            warnings.append(
                f"Date gap warning for {GOOD_WORKBOOK_DISPLAY_NAMES.get(market, market)}: "
                "one or more report_date gaps greater than 10 days were found in the source history."
            )

    return warnings


def _calculate_trader_master(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """Build the single source-of-truth calculated COT table.

    All derived fields are calculated here from raw/imported CFTC positioning
    columns. Dashboard, Market_Blocks and Raw_Data_Slim must use this output,
    not recalculate their own values.
    """
    required_value_cols = [
        "report_date",
        "market_name",
        "commercial_long",
        "commercial_short",
        "noncommercial_long",
        "noncommercial_short",
    ]
    warnings: list[str] = []
    master = df.copy()

    for column in required_value_cols + ["open_interest", "exchange", "cftc_contract_market_code", "source_report", "asset_class"]:
        if column not in master.columns:
            master[column] = pd.NA

    master["market_name"] = master["market_name"].map(_canonical_market_label)
    master = master[master["market_name"].isin(GOOD_WORKBOOK_MARKET_ORDER)].copy()
    master["report_date"] = pd.to_datetime(master["report_date"], errors="coerce")

    numeric_inputs = ["open_interest", "commercial_long", "commercial_short", "noncommercial_long", "noncommercial_short"]
    for column in numeric_inputs:
        master[column] = pd.to_numeric(master[column], errors="coerce")

    # De-duplicate and then strictly sort before calculating any time-series fields.
    # This guarantees previous-row logic never crosses markets and never uses
    # the incoming CFTC file order.
    master, duplicate_warnings = _deduplicate_for_calculation(master)
    warnings.extend(duplicate_warnings)
    master = _sort_workbook_rows(master)

    # Raw calculations. Do not use CFTC-provided change fields and do not keep
    # stale/hardcoded fields from prior workbooks.
    master["commercial_net"] = master["commercial_long"] - master["commercial_short"]
    master["noncommercial_net"] = master["noncommercial_long"] - master["noncommercial_short"]

    # Calculate strictly inside each market after sorting. No cross-market
    # contamination is possible because groupby("market_name") is the only shift
    # boundary used here.
    group = master.groupby("market_name", sort=False, group_keys=False)
    master["weekly_change"] = group["commercial_net"].diff(1)
    master["four_week_change"] = group["commercial_net"].diff(4)
    master["mm_weekly_change"] = group["noncommercial_net"].diff(1)
    master["mm_four_week_change"] = group["noncommercial_net"].diff(4)
    warnings.extend(_validate_market_date_order(master))

    def bias_from_change(value: object) -> str:
        if pd.isna(value) or value == 0:
            return "Neutral"
        return "Bullish" if value > 0 else "Bearish"

    master["bias"] = master["weekly_change"].apply(bias_from_change)
    master = _calculate_cot_scores(master)

    # Warn if any required source values are missing; leave calculations blank.
    for market in GOOD_WORKBOOK_MARKET_ORDER:
        market_rows = master[master["market_name"] == market]
        if market_rows.empty:
            warnings.append(
                f"Required market missing from workbook output: {GOOD_WORKBOOK_DISPLAY_NAMES.get(market, market)}. "
                "No matching CFTC rows were found after the strict market filter."
            )
            continue
        for column in required_value_cols:
            if market_rows[column].isna().any():
                warnings.append(
                    f"Required value missing for {GOOD_WORKBOOK_DISPLAY_NAMES.get(market, market)}: {column}. "
                    "Blank cells were kept; no values were faked."
                )

    return master.reset_index(drop=True), warnings



def _cot_strength(score: object) -> str:
    if pd.isna(score):
        return "Weak"
    value = float(score)
    if value <= 3:
        return "Weak"
    if value <= 6:
        return "Moderate"
    if value <= 8:
        return "Strong"
    return "Very Strong"


def _cot_summary(cot_bias: str, score: float, row: pd.Series) -> str:
    c1 = row.get("mm_weekly_change")
    c4 = row.get("mm_four_week_change")
    m1 = row.get("mm_weekly_change")
    managed_net = row.get("noncommercial_net")

    if cot_bias == "Bullish":
        if pd.notna(c4) and c4 > 0 and pd.notna(managed_net) and managed_net < 0 and pd.notna(m1) and m1 < 0:
            return "Managed money is strengthening over 1W and 4W from a net-short base. Bullish positioning support."
        if score >= 7:
            return "Bullish COT pressure from managed money with supportive positioning conditions."
        return "Managed money improved over 1W. Bullish COT bias, but confirmation is limited."

    if cot_bias == "Bearish":
        if pd.notna(c4) and c4 < 0 and pd.notna(managed_net) and managed_net > 0 and pd.notna(m1) and m1 > 0:
            return "Managed money is weakening over 1W and 4W from a net-long base. Bearish positioning pressure."
        if score >= 7:
            return "Bearish COT pressure from managed money with supportive positioning conditions."
        return "Managed money weakened over 1W. Bearish COT bias, but confirmation is limited."

    return "Neutral managed-money 1W change. No COT score applied."


def _calculate_cot_scores(master: pd.DataFrame) -> pd.DataFrame:
    """Add strict rule-based /10 COT scoring columns to Trader_Report master.

    Bias is driven only by managed-money 1W change. Points are awarded only in
    the direction of that bias. Neutral rows receive 0 points.
    """
    scored = master.copy()

    def score_row(row: pd.Series) -> pd.Series:
        c1 = row.get("mm_weekly_change")
        c4 = row.get("mm_four_week_change")
        managed_net = row.get("noncommercial_net")
        m1 = row.get("mm_weekly_change")

        if pd.isna(c1) or c1 == 0:
            cot_bias = "Neutral"
            cot_score = 0
        elif c1 > 0:
            cot_bias = "Bullish"
            cot_score = 2
            if pd.notna(c4) and c4 > 0:
                cot_score += 2
            if pd.notna(managed_net) and managed_net < 0:
                cot_score += 2
            if pd.notna(m1) and m1 < 0:
                cot_score += 2
            if pd.notna(m1) and m1 < 0:
                cot_score += 2
        else:
            cot_bias = "Bearish"
            cot_score = 2
            if pd.notna(c4) and c4 < 0:
                cot_score += 2
            if pd.notna(managed_net) and managed_net > 0:
                cot_score += 2
            if pd.notna(m1) and m1 > 0:
                cot_score += 2
            if pd.notna(m1) and m1 > 0:
                cot_score += 2

        cot_score = min(int(cot_score), 10)
        return pd.Series(
            {
                "cot_bias": cot_bias,
                "cot_score": cot_score,
                "cot_strength": _cot_strength(cot_score),
                "cot_summary": _cot_summary(cot_bias, float(cot_score), row),
            }
        )

    score_df = scored.apply(score_row, axis=1)
    for column in score_df.columns:
        scored[column] = score_df[column]
    return scored

def _build_source_notes(source_url: str, extra_sources: list[str] | None = None, warnings: list[str] | None = None) -> pd.DataFrame:
    rows = [
        {"field": "template_rule", "value": "The good workbook market set is the source of truth. Only original commodity markets plus NASDAQ and S&P 500 are exported."},
        {"field": "primary_source", "value": source_url},
        {"field": "cme_futures_only_source", "value": CME_FUTURES_ONLY_URL},
        {"field": "financial_futures_only_source_template", "value": FINANCIAL_FUTURES_ONLY_URL_TEMPLATE},
        {"field": "legacy_futures_only_source_template", "value": LEGACY_FUTURES_ONLY_URL_TEMPLATE},
        {"field": "notes", "value": "Financial Futures Only historical compressed data is used to backfill NASDAQ and S&P 500 CME equity-index futures."},
        {"field": "market_blocks", "value": "Market_Blocks is formatted as grouped market sections to match the historical dashboard template."},
        {"field": "scoring", "value": "Rule-based COT Bias, COT Score, COT Strength, and COT Summary are calculated from Trader_Report only. No external AI APIs are used."},
    ]
    for item in extra_sources or []:
        rows.append({"field": "additional_source", "value": item})
    for warning in warnings or []:
        rows.append({"field": "warning", "value": warning})
    return pd.DataFrame(rows)


def _clean_excel_value(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.date()
    return value


def _display_market(value: object) -> str:
    canonical = _canonical_market_label(value)
    return GOOD_WORKBOOK_DISPLAY_NAMES.get(canonical, str(value).strip())


def _contract_name(row: pd.Series) -> str:
    market = str(row.get("market_name", "")).upper()
    if market == "NASDAQ":
        return "E-MINI NASDAQ 100 - CHICAGO MERCANTILE EXCHANGE"
    if market == "S&P 500":
        return "E-MINI S&P 500 - CHICAGO MERCANTILE EXCHANGE"
    raw_market = row.get("market_name", "")
    exchange = row.get("exchange", "")
    if not pd.isna(exchange) and str(exchange).strip():
        return f"{raw_market} - {exchange}"
    return str(raw_market)


def _data_source(row: pd.Series) -> str:
    value = row.get("source_report")
    return "" if pd.isna(value) else str(value)


def _prepare_dashboard_table(dashboard: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, row in dashboard.iterrows():
        rows.append(
            {
                "Market": _display_market(row["market_name"]),
                "Latest Date": _clean_excel_value(row["report_date"]),
                "Commercial Net": row.get("commercial_net"),
                "1W Change": row.get("weekly_change"),
                "4W Change": row.get("four_week_change"),
                "Managed Money Net": row.get("noncommercial_net"),
                "MM 1W Change": row.get("mm_weekly_change"),
                "Bias": row.get("bias"),
                "cot_bias": row.get("cot_bias"),
                "cot_score": row.get("cot_score"),
                "cot_strength": row.get("cot_strength"),
                "cot_summary": row.get("cot_summary"),
                "CFTC Contract": _contract_name(row),
                "Data Source": _data_source(row),
            }
        )
    return pd.DataFrame(rows)


def _prepare_trader_report(df: pd.DataFrame) -> pd.DataFrame:
    records = []
    for _, row in df.iterrows():
        records.append(
            {
                "Date": _clean_excel_value(row.get("report_date")),
                "Market": _display_market(row.get("market_name")),
                "CFTC Contract": _contract_name(row),
                "Open Interest": row.get("open_interest"),
                "Commercial Long": row.get("commercial_long"),
                "Commercial Short": row.get("commercial_short"),
                "Commercial Net": row.get("commercial_net"),
                "Commercial Net 1W Chg": row.get("weekly_change"),
                "Commercial Net 4W Chg": row.get("four_week_change"),
                "Managed Money Long": row.get("noncommercial_long"),
                "Managed Money Short": row.get("noncommercial_short"),
                "Managed Money Net": row.get("noncommercial_net"),
                "MM Net 1W Chg": row.get("mm_weekly_change"),
                "MM Net 4W Chg": row.get("mm_four_week_change"),
                "Bias": row.get("bias"),
                "cot_bias": row.get("cot_bias"),
                "cot_score": row.get("cot_score"),
                "cot_strength": row.get("cot_strength"),
                "cot_summary": row.get("cot_summary"),
            }
        )
    return pd.DataFrame(records, columns=TRADER_REPORT_COLUMNS)


def _prepare_data_checks(master: pd.DataFrame) -> pd.DataFrame:
    required = [
        "report_date",
        "commercial_long",
        "commercial_short",
        "noncommercial_long",
        "noncommercial_short",
        "commercial_net",
        "noncommercial_net",
        "weekly_change",
        "four_week_change",
        "mm_weekly_change",
        "bias",
        "cot_bias",
        "cot_score",
        "cot_strength",
        "cot_summary",
    ]
    rows = []
    for market in GOOD_WORKBOOK_MARKET_ORDER:
        subset = master[master["market_name"] == market].copy() if "market_name" in master.columns else pd.DataFrame()
        missing = []
        if subset.empty:
            missing = required
            completed = False
            dates_strictly_increasing = False
            first_date = None
            last_date = None
        else:
            for column in required:
                if column not in subset.columns or subset[column].isna().all():
                    missing.append(column)
            first = pd.to_datetime(subset["report_date"], errors="coerce").min()
            last = pd.to_datetime(subset["report_date"], errors="coerce").max()
            first_date = first.date() if pd.notna(first) else None
            last_date = last.date() if pd.notna(last) else None
            dates = pd.to_datetime(subset["report_date"], errors="coerce").dropna().reset_index(drop=True)
            dates_strictly_increasing = bool(len(dates) <= 1 or (dates.diff().dropna() > pd.Timedelta(0)).all())
            # First rows naturally have missing 1W/4W changes. Calculations are
            # complete when the raw inputs and net fields exist, and at least one
            # 1W/4W result can be produced where the history is long enough.
            calc_cols = ["commercial_net", "noncommercial_net", "weekly_change", "four_week_change", "mm_weekly_change", "bias", "cot_bias", "cot_score", "cot_strength", "cot_summary"]
            completed = all(col in subset.columns for col in calc_cols) and subset["commercial_net"].notna().any() and subset["noncommercial_net"].notna().any() and dates_strictly_increasing
            if len(subset) >= 2:
                completed = completed and subset["weekly_change"].notna().any() and subset["mm_weekly_change"].notna().any()
            if len(subset) >= 5:
                completed = completed and subset["four_week_change"].notna().any()
        rows.append(
            {
                "market_name": GOOD_WORKBOOK_DISPLAY_NAMES.get(market, market),
                "row_count": int(len(subset)),
                "first_date": first_date,
                "last_date": last_date,
                "missing_required_columns": ", ".join(missing),
                "dates_strictly_increasing": bool(dates_strictly_increasing) if not subset.empty else False,
                "calculations_completed": bool(completed),
            }
        )
    return pd.DataFrame(rows)


def _write_dataframe(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> None:
    for c_idx, column in enumerate(df.columns, start_col):
        ws.cell(row=start_row, column=c_idx, value=column)
    for r_offset, (_, row) in enumerate(df.iterrows(), start_row + 1):
        for c_idx, column in enumerate(df.columns, start_col):
            ws.cell(row=r_offset, column=c_idx, value=_clean_excel_value(row[column]))


def _set_sheet_default(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"


def _style_header_row(ws, row: int, start_col: int, end_col: int, fill: str = HEADER_NAVY) -> None:
    for cell in ws[row][start_col - 1:end_col]:
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color=GRID))


def _style_title_row(ws, row: int, start_col: int, end_col: int, title: str) -> None:
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col, value=title)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(bold=True, color=WHITE, size=13)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    for col in range(start_col + 1, end_col + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=NAVY)


def _apply_bias_and_delta_formatting(ws, min_row: int, max_row: int, headers: dict[str, int]) -> None:
    green_font = "008000"
    red_font = "C00000"
    for row in range(min_row, max_row + 1):
        for header in ["1W Change", "4W Change", "Commercial Net 1W Chg", "Commercial Net 4W Chg", "MM 1W Change", "MM Net 1W Chg", "MM Net 4W Chg"]:
            col = headers.get(header)
            if not col:
                continue
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                if cell.value > 0:
                    cell.fill = PatternFill("solid", fgColor=POS_FILL)
                    cell.font = Font(color=green_font)
                elif cell.value < 0:
                    cell.fill = PatternFill("solid", fgColor=NEG_FILL)
                    cell.font = Font(color=red_font)
        for bias_header in ["Bias", "COT Bias", "cot_bias"]:
            bias_col = headers.get(bias_header)
            if bias_col:
                bias_cell = ws.cell(row=row, column=bias_col)
                bias = str(bias_cell.value or "").upper()
                if "BULL" in bias or "BUYING" in bias:
                    bias_cell.fill = PatternFill("solid", fgColor=GREEN_FILL)
                    bias_cell.font = Font(bold=True, color="006100")
                elif "BEAR" in bias or "SELLING" in bias:
                    bias_cell.fill = PatternFill("solid", fgColor=RED_FILL)
                    bias_cell.font = Font(bold=True, color="9C0006")
                elif "NEUTRAL" in bias:
                    bias_cell.fill = PatternFill("solid", fgColor=YELLOW_FILL)
                elif "FIRST" in bias:
                    bias_cell.fill = PatternFill("solid", fgColor="F3F4F6")
        score_col = headers.get("COT Score") or headers.get("cot_score")
        if score_col:
            ws.cell(row=row, column=score_col).number_format = "0"
        strength_col = headers.get("COT Strength") or headers.get("cot_strength")
        if strength_col:
            strength_cell = ws.cell(row=row, column=strength_col)
            strength = str(strength_cell.value or "").upper()
            if strength == "WEAK":
                strength_cell.fill = PatternFill("solid", fgColor=STRENGTH_WEAK_FILL)
            elif strength == "MODERATE":
                strength_cell.fill = PatternFill("solid", fgColor=STRENGTH_MODERATE_FILL)
            elif strength == "STRONG":
                strength_cell.fill = PatternFill("solid", fgColor=STRENGTH_STRONG_FILL)
                strength_cell.font = Font(bold=True, color="006100")
            elif strength == "VERY STRONG":
                strength_cell.fill = PatternFill("solid", fgColor=STRENGTH_VERY_STRONG_FILL)
                strength_cell.font = Font(bold=True, color=WHITE)


def _format_tabular_sheet(ws, max_row: int, max_col: int) -> None:
    _set_sheet_default(ws)
    _style_header_row(ws, 1, 1, max_col)
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    widths = {
        1: 12, 2: 18, 3: 44, 4: 14, 5: 16, 6: 16, 7: 16, 8: 18, 9: 18,
        10: 18, 11: 18, 12: 18, 13: 16, 14: 16, 15: 18, 16: 16, 17: 12, 18: 15, 19: 70,
    }
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 14)
    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = Border(bottom=Side(style="thin", color=GRID))
            if cell.column == 1:
                cell.number_format = "yyyy-mm-dd"
            elif isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
    headers = {ws.cell(row=1, column=col).value: col for col in range(1, max_col + 1)}
    _apply_bias_and_delta_formatting(ws, 2, max_row, headers)


def _format_dashboard(ws, dashboard_table: pd.DataFrame, all_rows: pd.DataFrame) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"
    end_col = len(dashboard_table.columns)
    end_letter = get_column_letter(end_col)
    ws.merge_cells(f"A1:{end_letter}1")
    ws["A1"] = "Historical COT Positioning Dashboard"
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = Font(bold=True, color=WHITE, size=16)
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.merge_cells(f"A2:{end_letter}2")
    ws["A2"] = "Built from the COT report template style + latest CFTC cleaned data"
    ws["A2"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A2"].font = Font(color=WHITE, italic=True)

    min_date = pd.to_datetime(all_rows["report_date"], errors="coerce").min()
    max_date = pd.to_datetime(all_rows["report_date"], errors="coerce").max()
    ws["A3"] = "Coverage"
    ws["B3"] = f"{min_date.date()} to {max_date.date()}" if pd.notna(min_date) and pd.notna(max_date) else ""
    ws["C3"] = "Rows"
    ws["D3"] = len(all_rows)
    ws["E3"] = "Markets populated"
    ws["F3"] = all_rows["market_name"].nunique()
    ws["A4"] = "Important limitation"
    ws["B4"] = "Dashboard is latest row per market only. Trader_Report and Market_Blocks contain the full available history. COT scoring is rule-based; no alerts or AI summaries are included yet."
    ws.merge_cells(f"B4:{end_letter}4")
    for cell in ws[3] + ws[4]:
        cell.alignment = Alignment(vertical="center")
    for cell in [ws["A3"], ws["C3"], ws["E3"], ws["A4"]]:
        cell.font = Font(bold=True)

    _write_dataframe(ws, dashboard_table, start_row=6, start_col=1)
    max_row = 6 + len(dashboard_table)
    _style_header_row(ws, 6, 1, end_col)
    ws.auto_filter.ref = f"A6:{get_column_letter(end_col)}{max_row}"
    widths = [18, 13, 16, 12, 12, 18, 14, 18, 16, 12, 15, 70, 42, 28]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in range(7, max_row + 1):
        ws.cell(row=row, column=2).number_format = "yyyy-mm-dd"
        for col in range(3, 8):
            ws.cell(row=row, column=col).number_format = "#,##0"
        for col in range(1, end_col + 1):
            ws.cell(row=row, column=col).border = Border(bottom=Side(style="thin", color=GRID))
    headers = {ws.cell(row=6, column=col).value: col for col in range(1, end_col + 1)}
    _apply_bias_and_delta_formatting(ws, 7, max_row, headers)


def _write_market_blocks(ws, trader_report: pd.DataFrame) -> None:
    """Render Market_Blocks from Trader_Report only.

    Trader_Report is the master calculated table. This function only formats a
    view of that table and never calculates COT metrics independently.
    """
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    current_row = 1

    report = trader_report.copy()
    report["_canonical_market"] = report["Market"].map(_canonical_market_label)
    report["Date"] = pd.to_datetime(report["Date"], errors="coerce")

    for market in GOOD_WORKBOOK_MARKET_ORDER:
        group = report[report["_canonical_market"] == market].sort_values("Date")
        if group.empty:
            continue

        _style_title_row(ws, current_row, 1, len(MARKET_BLOCK_COLUMNS), GOOD_WORKBOOK_DISPLAY_NAMES.get(market, market).upper())
        current_row += 1
        for col, header in enumerate(MARKET_BLOCK_COLUMNS, 1):
            ws.cell(row=current_row, column=col, value=header)
        _style_header_row(ws, current_row, 1, len(MARKET_BLOCK_COLUMNS))
        header_row = current_row
        current_row += 1

        for _, row in group.iterrows():
            values = [
                _clean_excel_value(row.get("Date")),
                row.get("Commercial Long"),
                row.get("Commercial Short"),
                row.get("Commercial Net"),
                row.get("Commercial Net 1W Chg"),
                row.get("Commercial Net 4W Chg"),
                row.get("Managed Money Net"),
                row.get("MM Net 1W Chg"),
                row.get("Bias"),
                row.get("cot_bias"),
                row.get("cot_score"),
                row.get("cot_strength"),
                row.get("cot_summary"),
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col, value=_clean_excel_value(value))
                cell.border = Border(bottom=Side(style="thin", color=GRID))
                if col == 1:
                    cell.number_format = "yyyy-mm-dd"
                elif col in {2, 3, 4, 5, 6, 7, 8}:
                    cell.number_format = "#,##0"
                elif col == 11:
                    cell.number_format = "0"
            current_row += 1

        headers = {ws.cell(row=header_row, column=col).value: col for col in range(1, len(MARKET_BLOCK_COLUMNS) + 1)}
        _apply_bias_and_delta_formatting(ws, header_row + 1, current_row - 1, headers)
        current_row += 1

    widths = [13, 18, 18, 18, 12, 12, 18, 13, 18, 16, 12, 15, 70]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _format_source_notes(ws, max_row: int) -> None:
    _set_sheet_default(ws)
    _style_header_row(ws, 1, 1, 2)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 115
    for row in range(2, max_row + 1):
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = Border(bottom=Side(style="thin", color=GRID))
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")


def _apply_workbook_formatting(workbook_path: Path, dashboard_table: pd.DataFrame, all_rows: pd.DataFrame, trader_report: pd.DataFrame) -> None:
    workbook = load_workbook(workbook_path)

    # Dashboard: rebuild into the original template-style title + latest table.
    ws = workbook["Dashboard"]
    ws.delete_rows(1, ws.max_row)
    _format_dashboard(ws, dashboard_table, all_rows)

    # Trader_Report: styled full-history table.
    ws = workbook["Trader_Report"]
    _format_tabular_sheet(ws, ws.max_row, ws.max_column)

    # Market_Blocks: rebuild into grouped market sections, exactly like the reference workbook.
    ws = workbook["Market_Blocks"]
    ws.delete_rows(1, ws.max_row)
    _write_market_blocks(ws, trader_report)

    ws = workbook["Raw_Data_Slim"]
    _format_tabular_sheet(ws, ws.max_row, ws.max_column)

    ws = workbook["Source_Notes"]
    _format_source_notes(ws, ws.max_row)

    ws = workbook["Data_Checks"]
    _format_tabular_sheet(ws, ws.max_row, ws.max_column)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 24

    workbook.save(workbook_path)



def _required_market_warnings(all_rows: pd.DataFrame) -> list[str]:
    warnings: list[str] = []
    if all_rows.empty or "market_name" not in all_rows.columns:
        return ["No rows were available for the required COT workbook market set."]
    canonical = all_rows["market_name"].map(_canonical_market_label)
    for required in GOOD_WORKBOOK_MARKET_ORDER:
        rows = all_rows[canonical == required]
        if rows.empty:
            warnings.append(f"Required market missing from workbook output: {GOOD_WORKBOOK_DISPLAY_NAMES.get(required, required)}. No matching CFTC rows were found after the strict market filter.")
            continue
        numeric_cols = ["commercial_long", "commercial_short", "commercial_net", "noncommercial_net"]
        existing = [col for col in numeric_cols if col in rows.columns]
        if existing and rows[existing].notna().sum().sum() == 0:
            warnings.append(f"Required market has rows but no populated positioning numbers: {GOOD_WORKBOOK_DISPLAY_NAMES.get(required, required)}. Check the CFTC source field mapping for this report family.")
    return warnings

def export_cot_workbook(
    df: pd.DataFrame,
    settings: Settings,
    source_url: str,
    dashboard_df: pd.DataFrame | None = None,
    extra_sources: list[str] | None = None,
    warnings: list[str] | None = None,
) -> ExportResult:
    ensure_dir(settings.exports_dir)
    ensure_dir(settings.processed_dir)

    run_date = date.today().isoformat()
    export_path = settings.exports_dir / f"cot_update_{run_date}.xlsx"
    processed_csv_path = settings.processed_dir / f"cot_cleaned_{run_date}.csv"

    df.to_csv(processed_csv_path, index=False)
    markets = markets_included(df)

    all_rows = _normalise_dashboard_input(dashboard_df if dashboard_df is not None else df)
    master_rows, calculation_warnings = _calculate_trader_master(all_rows)
    combined_warnings = list(warnings or []) + calculation_warnings + _required_market_warnings(master_rows)

    # Dashboard is latest-only per market. Trader_Report is the single master
    # calculated table. Market_Blocks is only a formatted view of Trader_Report.
    dashboard = (
        master_rows.sort_values(["market_name", "report_date"])
        .groupby("market_name", as_index=False, dropna=False)
        .tail(1)
        .reset_index(drop=True)
    )
    dashboard = _sort_workbook_rows(dashboard)
    dashboard_table = _prepare_dashboard_table(dashboard)
    trader_report = _prepare_trader_report(master_rows)
    raw_data_slim = master_rows[DASHBOARD_COLUMNS + ["mm_weekly_change", "mm_four_week_change", "cot_bias", "cot_score", "cot_strength", "cot_summary", "source_report"]].copy()
    data_checks = _prepare_data_checks(master_rows)
    source_notes = _build_source_notes(source_url, extra_sources=extra_sources, warnings=combined_warnings)

    with pd.ExcelWriter(export_path, engine="openpyxl") as writer:
        dashboard_table.to_excel(writer, sheet_name="Dashboard", index=False)
        trader_report.to_excel(writer, sheet_name="Trader_Report", index=False)
        # Temporary seed; overwritten by _write_market_blocks from Trader_Report.
        trader_report.to_excel(writer, sheet_name="Market_Blocks", index=False)
        raw_data_slim.to_excel(writer, sheet_name="Raw_Data_Slim", index=False)
        source_notes.to_excel(writer, sheet_name="Source_Notes", index=False)
        data_checks.to_excel(writer, sheet_name="Data_Checks", index=False)

    _apply_workbook_formatting(export_path, dashboard_table, master_rows, trader_report)

    return ExportResult(
        export_file_path=export_path,
        processed_csv_path=processed_csv_path,
        rows_exported=len(df),
        markets=markets,
    )
