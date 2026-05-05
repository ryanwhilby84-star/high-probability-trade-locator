from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


def format_table_sheet(ws: Worksheet) -> None:
    """Apply consistent readable formatting to a worksheet with headers in row 1."""
    if ws.max_row < 1 or ws.max_column < 1:
        return

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        header = str(column_cells[0].value or "")
        max_len = len(header)
        for cell in column_cells[1:]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 45)
